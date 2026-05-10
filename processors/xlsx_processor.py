from typing import Dict, List
from pathlib import Path
from .base import BaseProcessor

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XlsxProcessor(BaseProcessor):
    """Excel文件处理器 - 转换为Markdown表格格式"""

    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xls']

    def extract_text(self, file_path: str, to_markdown: bool = True) -> str:
        """从Excel提取文本，转换为Markdown表格格式"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("请安装openpyxl: pip install openpyxl")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            if to_markdown:
                return self._extract_as_markdown(wb)
            else:
                return self._extract_as_text(wb)
        except Exception as e:
            raise RuntimeError(f"XLSX解析失败: {str(e)}")

    def _extract_as_markdown(self, wb) -> str:
        """转换为Markdown表格 - 掐头去尾"""
        markdown_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            markdown_parts.append(f"## {sheet_name}")
            markdown_parts.append(self._to_markdown_table(rows))
            markdown_parts.append('')

        wb.close()
        return '\n'.join(markdown_parts)

    def _to_markdown_table(self, rows: List[tuple], max_rows: int = 50) -> str:
        """将行数据转换为Markdown表格 - 掐头去尾"""
        if not rows:
            return ""

        total_rows = len(rows)

        if total_rows <= max_rows:
            selected_rows = rows
        else:
            head_count = int(max_rows * 0.6)
            tail_count = int(max_rows * 0.3)
            selected_rows = rows[:head_count] + [('...',) * len(rows[0])] + rows[-tail_count:]

        markdown_lines = []

        for idx, row in enumerate(selected_rows):
            row_values = [str(cell) if cell is not None else '' for cell in row]
            line = '| ' + ' | '.join(row_values) + ' |'
            markdown_lines.append(line)

            if idx == 0:
                separator = '| ' + ' | '.join(['---'] * len(row_values)) + ' |'
                markdown_lines.append(separator)

        if total_rows > max_rows:
            markdown_lines.append(f'\n*共{total_rows}行，已省略中间{total_rows - max_rows}行*')

        return '\n'.join(markdown_lines)

    def _extract_as_text(self, wb) -> str:
        """提取为纯文本"""
        text_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"[工作表: {sheet_name}]")

            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else '' for cell in row]
                row_text = ' | '.join(v for v in row_values if v)
                if row_text:
                    text_parts.append(row_text)

            text_parts.append('')

        wb.close()
        return '\n'.join(text_parts)

    def extract_metadata(self, file_path: str) -> Dict:
        """提取Excel元数据"""
        info = self.get_file_info(file_path)
        metadata = {
            **info,
            'type': 'excel',
            'sheet_count': 0,
            'sheet_names': [],
            'row_count': 0
        }

        if not OPENPYXL_AVAILABLE:
            return metadata

        try:
            wb = load_workbook(file_path, read_only=True)
            metadata['sheet_count'] = len(wb.sheetnames)
            metadata['sheet_names'] = wb.sheetnames

            if wb.sheetnames:
                first_sheet = wb[wb.sheetnames[0]]
                metadata['row_count'] = sum(1 for _ in first_sheet.rows)

            wb.close()
        except Exception:
            pass

        return metadata

    def extract_rows(self, file_path: str, sheet_index: int = 0) -> List[Dict]:
        """提取Excel行数据为字典列表 - 数据集3专用，不转markdown"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("请安装openpyxl: pip install openpyxl")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                wb.close()
                return []

            headers = rows[0]
            data_rows = rows[1:]

            result = []
            for row in data_rows:
                row_dict = {}
                for idx, value in enumerate(row):
                    header = headers[idx] if idx < len(headers) else f'列{idx}'
                    row_dict[str(header)] = value
                result.append(row_dict)

            wb.close()
            return result
        except Exception as e:
            raise RuntimeError(f"XLSX行提取失败: {str(e)}")
